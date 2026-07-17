from pathlib import Path
from docx import Document 
from openpyxl import load_workbook
from pypdf import PdfReader

def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore")

def read_docx(path: Path) -> str:
    return "\n".join(p.text for p in Document(path).paragraphs)

def read_xlsx(path: Path) -> str:
    wb = load_workbook(path, data_only=True)
    out = []
    for sheet in wb.worksheets:
        out.append(f"# Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            out.append("\t".join("" if c is None else str(c) for c in row))
    return "\n".join(out)

def read_pdf(path: Path) -> str:
    return "\n".join(page.extract_text() or "" for page in PdfReader(path.pages))

READERS = {
    ".txt": read_text,
    ".py": read_text,
    ".docx": read_docx,
    ".xlsx": read_xlsx,
    ".pdf": read_pdf
}

def read_any(path: Path) -> str:
    reader = READERS.get(path.suffix.lower())
    if reader is None:
        raise ValueError(f"Unsupported file type: {path.suffix}")
    return reader(path)